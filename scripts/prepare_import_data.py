#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据清洗脚本（配置驱动版本）
使用列名而非列索引，避免硬编码
"""
import argparse
import yaml
from pathlib import Path
import openpyxl
from datetime import datetime
import re


class DataCleaner:
    """数据清洗器"""
    
    def __init__(self, config_path: str):
        """初始化清洗器"""
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)
    
    def clean_file(self, input_file: str, output_file: str = None, module: str = 'procurement'):
        """
        清洗数据文件
        
        Args:
            input_file: 输入文件路径
            output_file: 输出文件路径
            module: 模块名称
        """
        if output_file is None:
            output_file = self.config['output']['default_path']
        
        print(f"开始清洗数据...")
        print(f"  输入文件: {input_file}")
        print(f"  输出文件: {output_file}")
        print(f"  模块: {module}")
        
        # 读取Excel
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        # 获取列映射
        column_mapping = self.config['column_mapping'].get(module, {})
        
        if not column_mapping:
            print(f"警告: 未找到模块 {module} 的列映射配置")
            return
        
        # 找到列索引
        header_row = 1
        col_indices = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            col_name = str(cell.value).strip() if cell.value else ""
            if col_name in column_mapping:
                col_indices[column_mapping[col_name]] = col_idx
        
        print(f"  找到 {len(col_indices)} 个映射列")
        
        # 处理数据行
        cleaned_count = 0
        for row_idx in range(2, ws.max_row + 1):
            for field_name, col_idx in col_indices.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 应用清洗规则
                cleaned_value = self._apply_rules(cell.value, field_name)
                if cleaned_value != cell.value:
                    cell.value = cleaned_value
                    cleaned_count += 1
        
        # 保存
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        print(f"✓ 数据清洗完成: {output_file}")
        print(f"  清洗了 {cleaned_count} 个单元格")
    
    def _apply_rules(self, value, field_name):
        """应用清洗规则"""
        if value is None or value == "":
            return None
        
        # 查找字段规则
        rules = None
        for rule_config in self.config['cleanup_rules']:
            if rule_config['field'] == field_name:
                rules = rule_config['rules']
                break
        
        if not rules:
            return value
        
        # 依次应用规则
        result = value
        for rule in rules:
            rule_type = rule['type']
            
            try:
                if rule_type == 'trim':
                    result = str(result).strip()
                
                elif rule_type == 'remove_non_numeric':
                    result = re.sub(r'[^\d.-]', '', str(result))
                
                elif rule_type == 'convert_to_decimal':
                    try:
                        result = float(result) if result else None
                    except (ValueError, TypeError):
                        result = None
                
                elif rule_type == 'parse_date':
                    if isinstance(result, datetime):
                        result = result.date()
                    else:
                        try:
                            result = datetime.strptime(str(result), rule.get('format', '%Y-%m-%d')).date()
                        except (ValueError, TypeError):
                            result = None
                
                elif rule_type == 'validate_date_range':
                    if result and hasattr(result, 'year'):
                        min_year = rule.get('min_year', 2000)
                        max_year = rule.get('max_year', 2030)
                        if result.year < min_year or result.year > max_year:
                            result = None
                
                elif rule_type == 'standardize_choices':
                    mapping = rule.get('mapping', {})
                    result = mapping.get(str(result), result)
                    
            except Exception as e:
                print(f"警告: 应用规则 {rule_type} 到字段 {field_name} 失败: {e}")
        
        return result


def main():
    parser = argparse.ArgumentParser(description='数据清洗工具（配置驱动）')
    parser.add_argument('input_file', help='输入Excel文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径')
    parser.add_argument('--module', '-m', default='procurement', 
                       choices=['procurement', 'contract', 'payment'],
                       help='模块名称')
    parser.add_argument('--config', '-c', help='配置文件路径')
    
    args = parser.parse_args()
    
    # 默认配置路径
    if args.config is None:
        config_path = Path(__file__).parent / 'configs' / 'data_cleanup.yml'
    else:
        config_path = Path(args.config)
    
    if not config_path.exists():
        print(f"错误: 配置文件不存在: {config_path}")
        return 1
    
    # 执行清洗
    try:
        cleaner = DataCleaner(config_path)
        cleaner.clean_file(args.input_file, args.output, args.module)
        return 0
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit(main())