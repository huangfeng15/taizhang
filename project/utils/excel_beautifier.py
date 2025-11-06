"""
Excel工作表美化辅助函数
用于美化pandas生成的Excel工作表
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def beautify_worksheet(worksheet, money_columns=None, date_columns=None, center_columns=None):
    """
    美化Excel工作表
    
    Args:
        worksheet: openpyxl工作表对象
        money_columns: 金额列索引列表（从1开始），如 [5, 6, 7]
        date_columns: 日期列索引列表（从1开始），如 [8, 9]
        center_columns: 居中列索引列表（从1开始），如 [1, 2]
    """
    money_columns = money_columns or []
    date_columns = date_columns or []
    center_columns = center_columns or []
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='微软雅黑', size=10)
    money_font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # 美化表头（第1行）
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    worksheet.row_dimensions[1].height = 25
    
    # 美化数据行
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            # 应用边框
            cell.border = thin_border
            
            # 根据列类型应用不同格式
            if col_idx in money_columns:
                # 金额列：红色加粗，千分位分隔，不换行
                cell.font = money_font
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
            elif col_idx in date_columns:
                # 日期列：居中，不换行
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            elif col_idx in center_columns:
                # 居中列，自动换行
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # 默认：左对齐，自动换行
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 设置行高（增加高度以适应换行）
        worksheet.row_dimensions[row_idx].height = 30
    
    # 自动调整列宽
    for col_idx, column in enumerate(worksheet.columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for cell in column:
            try:
                if cell.value:
                    # 中文字符按2个字符宽度计算
                    cell_str = str(cell.value)
                    chinese_count = sum(1 for c in cell_str if ord(c) > 127)
                    cell_length = len(cell_str) + chinese_count
                    max_length = max(max_length, cell_length)
            except:
                pass
        
        # 设置列宽，限制在10-50之间
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'