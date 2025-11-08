"""
Excel导出美化工具
提供统一的Excel格式化和样式设置功能
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO


class ExcelExporter:
    """
    Excel导出美化工具类
    遵循KISS原则 - 提供简单易用的Excel格式化接口
    """
    
    # 预定义样式常量
    HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    CELL_FONT = Font(name='微软雅黑', size=10)
    CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    CELL_ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    CELL_ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    NUMBER_FONT = Font(name='微软雅黑', size=10, color='000000')
    MONEY_FONT = Font(name='微软雅黑', size=10, bold=True, color='C00000')
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # 默认行高和列宽
    DEFAULT_ROW_HEIGHT = 30
    HEADER_ROW_HEIGHT = 25
    MIN_COLUMN_WIDTH = 10
    MAX_COLUMN_WIDTH = 50
    
    def __init__(self, title='数据导出', sheet_name='Sheet1'):
        """
        初始化Excel导出器
        
        Args:
            title: Excel文件标题（用于文件名）
            sheet_name: 工作表名称
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.title = title
        
    def write_data(self, headers, data, column_formats=None):
        """
        写入数据并应用格式
        
        Args:
            headers: 表头列表 ['列名1', '列名2', ...]
            data: 数据列表 [[值1, 值2, ...], [值1, 值2, ...]]
            column_formats: 列格式字典 {列索引: 格式类型}
                格式类型: 'money'(金额), 'percent'(百分比), 'date'(日期), 
                         'center'(居中), 'right'(右对齐), 'number'(数字)
        """
        if not headers or not data:
            return
            
        column_formats = column_formats or {}
        
        # 写入表头
        self._write_headers(headers)
        
        # 写入数据行
        self._write_data_rows(data, column_formats)
        
        # 自动调整列宽
        self._auto_adjust_column_width(headers, data)
        
        # 冻结首行
        self.worksheet.freeze_panes = 'A2'
        
    def _write_headers(self, headers):
        """写入并格式化表头"""
        for col_idx, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 设置表头行高
        self.worksheet.row_dimensions[1].height = self.HEADER_ROW_HEIGHT
        
    def _write_data_rows(self, data, column_formats):
        """写入并格式化数据行"""
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.worksheet.cell(row=row_idx, column=col_idx)
                
                # 处理不同类型的值
                if value is None:
                    cell.value = ''
                elif isinstance(value, (date, datetime)):
                    cell.value = value.strftime('%Y-%m-%d') if isinstance(value, date) else value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    cell.value = '是' if value else '否'
                elif isinstance(value, Decimal):
                    cell.value = float(value)
                else:
                    cell.value = value
                
                # 应用格式
                format_type = column_formats.get(col_idx - 1)  # 列索引从0开始
                self._apply_cell_format(cell, format_type, value)
                
                # 应用边框
                cell.border = self.THIN_BORDER
            
            # 设置行高
            self.worksheet.row_dimensions[row_idx].height = self.DEFAULT_ROW_HEIGHT
    
    def _apply_cell_format(self, cell, format_type, value):
        """应用单元格格式"""
        if format_type == 'money':
            # 金额格式：红色加粗，千分位分隔，两位小数，不换行
            cell.font = self.MONEY_FONT
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
        elif format_type == 'number':
            # 普通数字格式：千分位分隔，两位小数，不换行
            cell.font = self.NUMBER_FONT
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
        elif format_type == 'percent':
            # 百分比格式，不换行
            cell.font = self.CELL_FONT
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        elif format_type == 'date':
            # 日期格式，不换行
            cell.font = self.CELL_FONT
            cell.number_format = 'yyyy-mm-dd'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        elif format_type == 'center':
            # 居中对齐，自动换行
            cell.font = self.CELL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif format_type == 'right':
            # 右对齐，自动换行
            cell.font = self.CELL_FONT
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        else:
            # 默认格式：左对齐，自动换行
            cell.font = self.CELL_FONT
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def _auto_adjust_column_width(self, headers, data):
        """自动调整列宽以适应内容"""
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            
            # 计算表头宽度
            max_length = len(str(header))
            
            # 计算数据列的最大宽度
            for row_data in data:
                if col_idx - 1 < len(row_data):
                    cell_value = row_data[col_idx - 1]
                    if cell_value is not None:
                        # 中文字符按2个字符宽度计算
                        cell_str = str(cell_value)
                        chinese_count = sum(1 for c in cell_str if ord(c) > 127)
                        cell_length = len(cell_str) + chinese_count
                        max_length = max(max_length, cell_length)
            
            # 设置列宽，限制在最小和最大宽度之间
            adjusted_width = min(max(max_length + 2, self.MIN_COLUMN_WIDTH), self.MAX_COLUMN_WIDTH)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_response(self, filename=None):
        """
        保存为HTTP响应
        
        Args:
            filename: 文件名，如果为None则使用标题
            
        Returns:
            HttpResponse对象
        """
        if filename is None:
            filename = f"{self.title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # 保存到内存
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        # 创建响应
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def save_to_file(self, file_path):
        """
        保存到文件
        
        Args:
            file_path: 文件路径
        """
        self.workbook.save(file_path)
        return file_path


def export_data_to_excel(headers, data, title='数据导出', column_formats=None, filename=None):
    """
    便捷函数：快速导出数据为Excel
    
    Args:
        headers: 表头列表
        data: 数据列表
        title: 标题
        column_formats: 列格式字典
        filename: 文件名
        
    Returns:
        HttpResponse对象
        
    Example:
        headers = ['项目编码', '项目名称', '合同金额', '付款比例']
        data = [
            ['PRJ001', '项目A', 1000000.00, 0.75],
            ['PRJ002', '项目B', 2000000.00, 0.50],
        ]
        column_formats = {
            2: 'money',  # 合同金额列使用金额格式
            3: 'percent'  # 付款比例列使用百分比格式
        }
        return export_data_to_excel(headers, data, '项目统计', column_formats)
    """
    exporter = ExcelExporter(title=title)
    exporter.write_data(headers, data, column_formats)
    return exporter.save_to_response(filename)