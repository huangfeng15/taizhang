"""
统一的报表生成器
整合BaseReportGenerator, ProfessionalReportGenerator, AdvancedReportGenerator
遵循SOLID原则的重构版本
"""
from datetime import date, timedelta
from typing import Optional, List, Dict, Any
from django.utils import timezone

from project.services.report_data_service import ReportDataService
from project.services.word_formatter import WordFormatter


class ReportGenerator:
    """
    统一的报表生成器
    遵循SOLID原则：
    - SRP: 只负责协调数据服务和格式化器
    - OCP: 通过report_type参数扩展功能
    - DIP: 依赖抽象（ReportDataService）而非具体实现
    """

    def __init__(self, start_date: date, end_date: date, project_codes: Optional[List[str]] = None):
        """
        初始化报表生成器

        Args:
            start_date: 统计开始日期
            end_date: 统计结束日期
            project_codes: 项目编码列表，None表示全部项目
        """
        self.start_date = start_date
        self.end_date = end_date
        self.project_codes = project_codes

        # 依赖注入（DIP原则）
        self.data_service = ReportDataService(start_date, end_date, project_codes)

    def generate_word_report(self, report_type: str, file_path: str, report_title: str = None) -> str:
        """
        生成Word报表

        Args:
            report_type: 报表类型 ('standard' | 'professional' | 'comprehensive')
            file_path: 输出文件路径
            report_title: 报告标题（可选）

        Returns:
            str: 生成的文件路径
        """
        # 获取数据（SRP - 数据获取与格式化分离）
        report_data = self._prepare_report_data(report_type, report_title)

        # 格式化输出（OCP - 通过参数控制格式）
        formatter = WordFormatter(template_type=report_type)
        return formatter.format_report(report_data, file_path)

    def generate_data_only(self, report_type: str = 'standard', report_title: str = None) -> Dict[str, Any]:
        """
        仅生成报表数据，不生成Word文档
        用于API接口或其他格式的导出

        Args:
            report_type: 报表类型
            report_title: 报告标题

        Returns:
            dict: 报表数据
        """
        return self._prepare_report_data(report_type, report_title)

    def generate_data(self, report_type: str = 'standard', report_title: str = None) -> Dict[str, Any]:
        """
        向后兼容方法：保留 generate_data 接口，内部委托 generate_data_only。

        说明：
        - 旧代码和文档中大量使用 generator.generate_data(...) 调用形式
        - 为避免在多个调用点同时改动，这里提供一个轻量别名方法
        """
        return self.generate_data_only(report_type=report_type, report_title=report_title)

    def _prepare_report_data(self, report_type: str, report_title: str = None) -> Dict[str, Any]:
        """
        准备报表数据

        Args:
            report_type: 报表类型
            report_title: 报告标题

        Returns:
            dict: 完整的报表数据
        """
        # 确定报告标题
        if not report_title:
            report_title = self._get_default_title(report_type)

        # 基础数据（所有报表类型都需要）
        base_data = {
            'meta': self.data_service.get_report_meta(report_type, report_title),
            'summary': self.data_service.get_executive_summary(),
        }

        # 根据报表类型获取不同详细程度的数据
        if report_type == 'standard':
            base_data.update(self._get_standard_data())
        elif report_type == 'professional':
            base_data.update(self._get_professional_data())
        elif report_type == 'comprehensive':
            base_data.update(self._get_comprehensive_data())

        return base_data

    def _get_standard_data(self) -> Dict[str, Any]:
        """获取标准报表数据"""
        return self.data_service.get_all_statistics()

    def _get_professional_data(self) -> Dict[str, Any]:
        """获取专业报表数据"""
        data = self.data_service.get_all_statistics()

        # 添加专业报表特有的数据
        data['projects_overview'] = self.data_service.get_projects_overview()
        data['archive_monitoring'] = self.data_service.get_archive_monitoring()
        data['completeness'] = self.data_service.get_completeness_analysis()
        data['ranking'] = self.data_service.get_ranking_analysis()
        data['recommendations'] = self.data_service.get_recommendations()

        # 单项目报告增加项目特定信息
        if self.data_service.is_single_project:
            data['project_details'] = self.data_service.get_single_project_details()

        return data

    def _get_comprehensive_data(self) -> Dict[str, Any]:
        """获取综合报表数据"""
        data = self._get_professional_data()

        # 添加综合报表特有的深度分析数据
        data['financial_analysis'] = self.data_service.get_financial_analysis()

        # 重命名部分字段以匹配综合报表格式
        data['executive_summary'] = data.pop('summary')
        data['procurement_comprehensive'] = data.pop('procurement')
        data['contract_comprehensive'] = data.pop('contract')
        data['payment_comprehensive'] = data.pop('payment')
        data['settlement_comprehensive'] = data.pop('settlement')

        return data

    def _get_default_title(self, report_type: str) -> str:
        """获取默认报告标题"""
        title_map = {
            'standard': '工作报告',
            'professional': '专业工作报告',
            'comprehensive': '综合工作报告',
        }
        return title_map.get(report_type, '工作报告')


# ========== 便捷的报表生成器子类 ==========
# 保持向后兼容，提供与原有代码相同的接口

class WeeklyReportGenerator(ReportGenerator):
    """周报生成器"""

    def __init__(self, year: int, week: int, project_codes: Optional[List[str]] = None):
        """
        初始化周报生成器

        Args:
            year: 年份
            week: 周数（1-53）
            project_codes: 项目编码列表
        """
        # 计算周的起止日期
        start_date, end_date = self._calculate_week_range(year, week)
        super().__init__(start_date, end_date, project_codes)
        self.year = year
        self.week = week

    def _calculate_week_range(self, year: int, week: int) -> tuple:
        """计算周的起止日期"""
        # 获取该年第一天
        jan_1 = date(year, 1, 1)
        # 计算第一周的开始日期（周一）
        days_to_monday = (7 - jan_1.weekday()) % 7
        first_monday = jan_1 + timedelta(days=days_to_monday)
        # 计算目标周的开始日期
        start_date = first_monday + timedelta(weeks=week - 1)
        end_date = start_date + timedelta(days=6)
        return start_date, end_date

    def generate_report(self, file_path: str) -> str:
        """生成周报"""
        title = f"{self.year}年第{self.week}周工作报告"
        return self.generate_word_report('professional', file_path, title)


class MonthlyReportGenerator(ReportGenerator):
    """月报生成器"""

    def __init__(self, year: int, month: int, project_codes: Optional[List[str]] = None):
        """
        初始化月报生成器

        Args:
            year: 年份
            month: 月份（1-12）
            project_codes: 项目编码列表
        """
        # 计算月的起止日期
        start_date, end_date = self._calculate_month_range(year, month)
        super().__init__(start_date, end_date, project_codes)
        self.year = year
        self.month = month

    def _calculate_month_range(self, year: int, month: int) -> tuple:
        """计算月的起止日期"""
        start_date = date(year, month, 1)
        # 计算下个月的第一天，然后减一天得到本月最后一天
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        return start_date, end_date

    def generate_report(self, file_path: str) -> str:
        """生成月报"""
        title = f"{self.year}年{self.month}月工作报告"
        return self.generate_word_report('professional', file_path, title)


class QuarterlyReportGenerator(ReportGenerator):
    """季报生成器"""

    def __init__(self, year: int, quarter: int, project_codes: Optional[List[str]] = None):
        """
        初始化季报生成器

        Args:
            year: 年份
            quarter: 季度（1-4）
            project_codes: 项目编码列表
        """
        # 计算季度的起止日期
        start_date, end_date = self._calculate_quarter_range(year, quarter)
        super().__init__(start_date, end_date, project_codes)
        self.year = year
        self.quarter = quarter

    def _calculate_quarter_range(self, year: int, quarter: int) -> tuple:
        """计算季度的起止日期"""
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        start_date = date(year, start_month, 1)
        # 计算季度最后一天
        if end_month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, end_month + 1, 1) - timedelta(days=1)
        return start_date, end_date

    def generate_report(self, file_path: str) -> str:
        """生成季报"""
        title = f"{self.year}年第{self.quarter}季度工作报告"
        return self.generate_word_report('professional', file_path, title)


class AnnualReportGenerator(ReportGenerator):
    """年报生成器"""

    def __init__(self, year: int, project_codes: Optional[List[str]] = None):
        """
        初始化年报生成器

        Args:
            year: 年份
            project_codes: 项目编码列表
        """
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        super().__init__(start_date, end_date, project_codes)
        self.year = year

    def generate_report(self, file_path: str, report_type: str = 'comprehensive') -> str:
        """
        生成年报

        Args:
            file_path: 输出文件路径
            report_type: 报表类型，年报默认使用comprehensive

        Returns:
            str: 生成的文件路径
        """
        title = f"{self.year}年度工作报告"
        return self.generate_word_report(report_type, file_path, title)


# ========== 向后兼容的导出函数 ==========

def export_to_word_professional(report_data: Dict[str, Any], file_path: str) -> str:
    """
    导出专业Word文档格式的报告（向后兼容）

    Args:
        report_data: 报告数据字典
        file_path: 导出文件路径

    Returns:
        str: 导出的文件路径
    """
    formatter = WordFormatter(template_type='professional')
    return formatter.format_report(report_data, file_path)


def export_comprehensive_word_report(report_data: Dict[str, Any], file_path: str) -> str:
    """
    导出综合Word报告（向后兼容）

    Args:
        report_data: 报告数据（来自AdvancedReportGenerator）
        file_path: 导出文件路径

    Returns:
        str: 导出的文件路径
    """
    formatter = WordFormatter(template_type='comprehensive')
    return formatter.format_report(report_data, file_path)


def export_to_word(report_data: Dict[str, Any], file_path: str) -> str:
    """
    标准版 Word 导出，兼容旧版 report_generator.export_to_word 接口。

    Args:
        report_data: 报表数据字典
        file_path: 输出文件路径

    Returns:
        str: 生成的文件路径
    """
    # 优先从元信息中识别报告类型，以便按 standard/professional/comprehensive 选择模板
    meta = report_data.get('meta', {}) if isinstance(report_data, dict) else {}
    template_type = meta.get('report_type', 'standard')
    if template_type not in ('standard', 'professional', 'comprehensive'):
        template_type = 'standard'

    formatter = WordFormatter(template_type=template_type)
    return formatter.format_report(report_data, file_path)


def export_to_excel(report_data: Dict[str, Any], file_path: str) -> str:
    """
    Excel 导出，兼容旧版 report_generator.export_to_excel 接口。
    实现遵循 KISS：将顶层为列表[dict]的数据各写入独立工作表，
    其余标量/字典信息写入“Summary”表。

    注意：依赖 openpyxl，请确保环境已安装。
    """
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise ImportError("openpyxl 未安装或不可用，无法导出 Excel") from e

    def safe_sheet_name(name: str) -> str:
        # Excel 工作表名限制：最多31字符，且不能包含 \ / * ? : [ ]
        invalid = set('\\/*?:[]')
        cleaned = ''.join(ch for ch in str(name) if ch not in invalid)
        return (cleaned or 'Sheet')[:31]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    # 先写 meta 和 summary
    def write_kv(ws, start_row, data_dict):
        r = start_row
        for k, v in (data_dict or {}).items():
            ws.cell(row=r, column=1, value=str(k))
            ws.cell(row=r, column=2, value=str(v))
            r += 1
        return r

    row = 1
    if isinstance(report_data, dict):
        # 写 meta
        if isinstance(report_data.get('meta'), dict):
            ws_summary.cell(row=row, column=1, value='[meta]')
            row += 1
            row = write_kv(ws_summary, row, report_data.get('meta')) + 1

        # 写 summary / executive_summary
        summary_key = 'summary' if 'summary' in report_data else 'executive_summary'
        if isinstance(report_data.get(summary_key), dict):
            ws_summary.cell(row=row, column=1, value='[summary]')
            row += 1
            row = write_kv(ws_summary, row, report_data.get(summary_key)) + 1

        # 其余顶层 list[dict] -> 独立工作表
        for key, value in report_data.items():
            if key in ('meta', 'summary', 'executive_summary'):
                continue
            if isinstance(value, list) and value and isinstance(value[0], dict):
                ws = wb.create_sheet(title=safe_sheet_name(key))
                # 收集列
                headers = set()
                for item in value:
                    headers.update(item.keys())
                headers = list(headers)
                for j, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=j, value=str(h))
                # 写数据
                r = 2
                for item in value:
                    for j, h in enumerate(headers, start=1):
                        ws.cell(row=r, column=j, value=str(item.get(h, '')))
                    r += 1

    wb.save(file_path)
    return file_path
