"""
工作周期监控视图 V2.0
完全按照需求方案文档重构
"""
import json
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from project.views_helpers import _resolve_global_filters, _extract_monitoring_filters
from project.services.monitors.cycle_statistics_v2 import CycleStatisticsServiceV2
from project.services.monitors.config import CYCLE_RULES


@require_http_methods(["GET"])
def cycle_monitor_v2(request):
    """
    工作周期监控主页面
    
    支持功能：
    1. 总览统计卡片（采购周期、合同周期、数据质量）
    2. 趋势图（周/月）
    3. 维度拆解（采购方式等）
    4. 明细表
    5. 筛选与分组
    """
    # 解析全局筛选参数
    global_filters = _resolve_global_filters(request)
    year_context, project_codes, project_filter_config, filter_config = _extract_monitoring_filters(request)
    
    # 解析页面特定参数
    procurement_methods = request.GET.getlist('procurement_method')
    time_dimension = request.GET.get('time_dimension', 'requirement_approval')
    group_by = request.GET.get('group_by', '')
    show_overdue_only = request.GET.get('show_overdue_only', '') == 'true'
    export_format = request.GET.get('export', '')
    
    # 初始化服务
    service = CycleStatisticsServiceV2()
    
    # 获取总览统计
    overview_stats = service.get_overview_statistics(
        year_filter=global_filters['year_value'],
        project_filter=global_filters['project'],
        procurement_methods=procurement_methods,
        time_dimension=time_dimension
    )
    
    # 获取趋势数据
    trend_data = service.get_trend_data(
        year_filter=global_filters['year_value'],
        project_filter=global_filters['project'],
        procurement_methods=procurement_methods,
        time_dimension=time_dimension
    )
    
    # 获取维度拆解数据（如果请求了分组）
    dimension_breakdown = []
    if group_by == 'procurement_method':
        dimension_breakdown = service.get_dimension_breakdown(
            dimension='procurement_method',
            year_filter=global_filters['year_value'],
            project_filter=global_filters['project'],
            top_n=20
        )
    
    # 获取明细记录
    detail_records = service.get_detail_records(
        year_filter=global_filters['year_value'],
        project_filter=global_filters['project'],
        procurement_methods=procurement_methods,
        show_overdue_only=show_overdue_only,
        time_dimension=time_dimension
    )
    
    # 如果请求导出
    if export_format in ['csv', 'excel']:
        return _export_detail_records(detail_records, export_format)
    
    # 构建上下文
    context = {
        'page_title': '工作周期监控',
        'filter_config': filter_config,
        
        # 总览统计
        'overview_stats': overview_stats,
        
        # 趋势数据（JSON格式供前端图表使用）
        'trend_data_json': json.dumps(trend_data, ensure_ascii=False),
        
        # 维度拆解
        'dimension_breakdown': dimension_breakdown,
        'group_by': group_by,
        
        # 明细记录
        'detail_records': detail_records,
        'show_overdue_only': show_overdue_only,
        
        # 筛选参数回显
        'selected_procurement_methods': procurement_methods,
        'time_dimension': time_dimension,
        
        # 配置数据
        'cycle_rules': CYCLE_RULES,
        'procurement_method_choices': list(CYCLE_RULES['procurement']['deadline_map'].keys()),
        'time_dimension_choices': [
            ('requirement_approval', '需求审批完成时间'),
            ('result_publicity', '公示完成时间'),
            ('contract_signing', '合同签订时间')
        ],
        
        # SLA配置
        'sla_procurement_default': CYCLE_RULES['procurement']['default_deadline'],
        'sla_contract': CYCLE_RULES['contract']['deadline_days'],
    }
    
    return render(request, 'monitoring/cycle_v2.html', context)


@require_http_methods(["GET"])
def cycle_monitor_api(request):
    """
    工作周期监控API接口
    用于AJAX请求获取数据
    """
    action = request.GET.get('action', 'overview')
    
    # 解析参数
    global_filters = _resolve_global_filters(request)
    procurement_methods = request.GET.getlist('procurement_method')
    time_dimension = request.GET.get('time_dimension', 'requirement_approval')
    
    # 初始化服务
    service = CycleStatisticsServiceV2()
    
    try:
        if action == 'overview':
            # 获取总览统计
            data = service.get_overview_statistics(
                year_filter=global_filters['year_value'],
                project_filter=global_filters['project'],
                procurement_methods=procurement_methods,
                time_dimension=time_dimension
            )
            return JsonResponse({'success': True, 'data': data})
        
        elif action == 'trend':
            # 获取趋势数据
            data = service.get_trend_data(
                year_filter=global_filters['year_value'],
                project_filter=global_filters['project'],
                procurement_methods=procurement_methods,
                time_dimension=time_dimension
            )
            return JsonResponse({'success': True, 'data': data})
        
        elif action == 'breakdown':
            # 获取维度拆解
            dimension = request.GET.get('dimension', 'procurement_method')
            top_n = int(request.GET.get('top_n', 10))
            data = service.get_dimension_breakdown(
                dimension=dimension,
                year_filter=global_filters['year_value'],
                project_filter=global_filters['project'],
                top_n=top_n
            )
            return JsonResponse({'success': True, 'data': data})
        
        elif action == 'details':
            # 获取明细记录
            show_overdue_only = request.GET.get('show_overdue_only', '') == 'true'
            data = service.get_detail_records(
                year_filter=global_filters['year_value'],
                project_filter=global_filters['project'],
                procurement_methods=procurement_methods,
                show_overdue_only=show_overdue_only,
                time_dimension=time_dimension
            )
            return JsonResponse({'success': True, 'data': data})
        
        else:
            return JsonResponse({'success': False, 'message': f'未知的action: {action}'}, status=400)
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def _export_detail_records(records, format_type):
    """导出明细记录"""
    import csv
    from io import StringIO
    
    if format_type == 'csv':
        # 创建CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = [
            '项目ID', '项目名称', '标段/包',
            '需求审批完成时间', '结果公示完成时间', '合同签订日期',
            '采购周期(天)', '合同周期(天)',
            '采购方式', '金额', '币种', '含税',
            '业务部门', '项目经理/经办人', '供应商',
            '状态', '采购是否逾期', '合同是否逾期',
            '采购SLA阈值(天)', '合同SLA阈值(天)', '备注'
        ]
        writer.writerow(headers)
        
        # 写入数据
        for record in records:
            row = [
                record['project_id'],
                record['project_name'],
                record['section_id'],
                record['demand_approval_time'] or '',
                record['result_publish_time'] or '',
                record['contract_sign_date'] or '',
                record['procurement_cycle_days'] or '',
                record['contract_cycle_days'] or '',
                record['procurement_method'],
                record['amount'] or '',
                record['currency'],
                record['tax_included'],
                record['department'],
                record['owner'],
                record['supplier'],
                record['status'],
                '是' if record['is_overdue_proc'] else '否',
                '是' if record['is_overdue_cont'] else '否',
                record['sla_proc_days'],
                record['sla_cont_days'],
                record['remark']
            ]
            writer.writerow(row)
        
        # 返回响应
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="cycle_details.csv"'
        return response
    
    elif format_type == 'excel':
        # Excel导出（需要openpyxl库）
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "工作周期明细"
            
            # 写入表头
            headers = [
                '项目ID', '项目名称', '标段/包',
                '需求审批完成时间', '结果公示完成时间', '合同签订日期',
                '采购周期(天)', '合同周期(天)',
                '采购方式', '金额', '币种', '含税',
                '业务部门', '项目经理/经办人', '供应商',
                '状态', '采购是否逾期', '合同是否逾期',
                '采购SLA阈值(天)', '合同SLA阈值(天)', '备注'
            ]
            ws.append(headers)
            
            # 写入数据
            for record in records:
                row = [
                    record['project_id'],
                    record['project_name'],
                    record['section_id'],
                    str(record['demand_approval_time']) if record['demand_approval_time'] else '',
                    str(record['result_publish_time']) if record['result_publish_time'] else '',
                    str(record['contract_sign_date']) if record['contract_sign_date'] else '',
                    record['procurement_cycle_days'] or '',
                    record['contract_cycle_days'] or '',
                    record['procurement_method'],
                    float(record['amount']) if record['amount'] else '',
                    record['currency'],
                    record['tax_included'],
                    record['department'],
                    record['owner'],
                    record['supplier'],
                    record['status'],
                    '是' if record['is_overdue_proc'] else '否',
                    '是' if record['is_overdue_cont'] else '否',
                    record['sla_proc_days'],
                    record['sla_cont_days'],
                    record['remark']
                ]
                ws.append(row)
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # 保存到内存
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 返回响应
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="cycle_details.xlsx"'
            return response
        
        except ImportError:
            # 如果没有openpyxl，降级到CSV
            return _export_detail_records(records, 'csv')
    
    else:
        return HttpResponse('不支持的导出格式', status=400)