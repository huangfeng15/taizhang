"""
Excel 导入模板生成器
根据 YAML 配置文件生成标准化的导入模板
遵循 DRY 原则，避免硬编码
"""
import os
import yaml
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from django.apps import apps
from project.helptext import get_message
from project.constants import BASE_YEAR, get_current_year


def get_import_template_config():
    """
    获取导入模板配置定义
    避免循环导入问题
    """
    return {
        'project': {
            'long': {
                'filename': 'project_import_template_long.csv',
                'headers': [
                    '项目编码',
                    '序号',
                    '项目名称',
                    '项目描述',
                    '项目负责人',
                    '项目状态',
                    '备注',
                    '模板说明',
                ],
                'notes': [
                    '【必填字段】项目编码*、项目名称*（标记*号的为必填字段，不能为空）',
                    '【编码规则】项目编码仅允许字母、数字、中文、连字符(-)、下划线(_)和点(.)，禁止使用 / 等特殊字符',
                    '【状态选项】项目状态可选值：进行中、已完成、已暂停、已取消（留空默认为"进行中"）',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
        },
        'procurement': {
            'long': {
                'filename': 'procurement_import_template_long.csv',
                'headers': [
                    '项目编码',
                    '序号',
                    '招采编号',
                    '采购项目名称',
                    '采购单位',
                    '中标单位',
                    '中标单位联系人及方式',
                    '采购方式',
                    '采购类别',
                    '采购预算金额(元)',
                    '采购控制价（元）',
                    '中标金额（元）',
                    '计划结束采购时间',
                    '候选人公示结束时间',
                    '结果公示发布时间',
                    '中标通知书发放日期',
                    '采购经办人',
                    '需求部门',
                    '申请人联系电话（需求部门）',
                    '采购需求书审批完成日期（OA）',
                    '采购平台',
                    '资格审查方式',
                    '评标谈判方式',
                    '定标方法',
                    '公告发布时间',
                    '报名截止时间',
                    '开标时间',
                    '评标委员会成员',
                    '投标担保形式及金额（元）',
                    '投标担保退回日期',
                    '履约担保形式及金额（元）',
                    '候选人公示期质疑情况',
                    '应招未招说明（由公开转单一或邀请的情况）',
                    '资料归档日期',
                    '模板说明',
                ],
                'notes': [
                    '【必填字段】招采编号*、采购项目名称*（标记*号的为必填字段，不能为空）',
                    '【编码规则】招采编号仅允许字母、数字、中文、连字符(-)、下划线(_)和点(.)，禁止使用 / 等特殊字符。建议格式：GC2025001',
                    '【项目关联】项目编码字段用于关联已存在的项目，必须填写系统中已存在的项目编码',
                    '【时间要求】公告发布时间、报名截止时间、开标时间、候选人公示结束时间、结果公示发布时间等均使用 YYYY-MM-DD 格式',
                    '【金额格式】所有金额列仅填写数字（可带小数），单位为元，例如：1500000.00 或 1500000',
                    '【采购方式】常见选项：公开招标、单一来源采购、公开询价、直接采购、公开竞价、战采结果应用等，可结合采购类别填写',
                    '【担保信息】投标担保与履约担保可填写形式与金额，例如：银行保函 500000.00',
                    '【质疑情况】候选人公示期质疑情况用于记录公示期处理情况，可留空',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
        },
        'contract': {
            'long': {
                'filename': 'contract_import_template_long.csv',
                'headers': [
                    '项目编码',
                    '关联采购编号',
                    '文件定位',
                    '合同来源',
                    '关联主合同编号',
                    '序号',
                    '合同序号',
                    '合同编号',
                    '合同名称',
                    '合同签订经办人',
                    '合同类型',
                    '甲方',
                    '乙方',
                    '含税签约合同价（元）',
                    '合同签订日期',
                    '甲方法定代表人及联系方式',
                    '甲方联系人及联系方式',
                    '甲方负责人及联系方式',
                    '乙方法定代表人及联系方式',
                    '乙方联系人及联系方式',
                    '乙方负责人及联系方式',
                    '合同工期/服务期限',
                    '支付方式',
                    '履约担保退回时间',
                    '资料归档日期',
                    '模板说明',
                ],
                'notes': [
                    '【必填字段】合同编号*、合同名称*、甲方*、乙方*、合同签订日期*（标记*号的为必填字段，不能为空）',
                    '【编码规则】合同编号与合同序号仅允许字母、数字、中文、连字符(-)、下划线(_)和点(.)，禁止使用 / 等特殊字符',
                    '【文件定位】仅支持四种类型：主合同、补充协议、解除协议、框架协议（留空默认为"主合同"）',
                    '【关联规则】补充协议或解除协议必须填写"关联主合同编号"，关联已存在的主合同',
                    '【合同类型】第11列的合同类型用于描述合同性质，如服务类、货物类、工程类等',
                    '【合同来源】可选值：采购合同、直接签订（留空默认为"采购合同"）',
                    '【项目关联】项目编码字段用于关联已存在的项目，必须填写系统中已存在的项目编码',
                    '【采购关联】关联采购编号字段用于关联已存在的采购记录，如无采购可留空',
                    '【联系信息】甲乙方的法定代表人、联系人、负责人信息均为可选字段，建议填写完整以便管理',
                    '【日期格式】合同签订日期、履约担保退回时间等日期字段统一使用 YYYY-MM-DD 格式',
                    '【金额格式】含税签约合同价仅填写数字（可带小数），单位为元，例如：2500000.00',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
        },
        'payment': {
            'long': {
                'filename': 'payment_import_template_long.csv',
                'headers': [
                    '项目编码',
                    '序号',
                    '付款编号',
                    '关联合同编号',
                    '实付金额(元)',
                    '付款日期',
                    '结算价（元）',
                    '是否办理结算',
                    '模板说明',
                ],
                'notes': [
                    '【必填字段】关联合同编号*、实付金额*、付款日期*（标记*号的为必填字段，不能为空）',
                    '【编码规则】付款编号可留空由系统自动生成；如手动填写需遵守编号格式限制（禁止 / 等特殊字符）',
                    '【合同关联】关联合同编号必须填写系统中已存在的合同编号或合同序号',
                    '【日期格式】付款日期必须使用 YYYY-MM-DD 格式，例如：2025-10-20',
                    '【金额格式】实付金额和结算价仅填写数字（可带小数），单位为元，例如：500000.00',
                    '【结算标记】是否办理结算可填写：是、否、true、false（留空默认为"否"）',
                    '【结算价说明】如果该笔付款是结算付款，需在"结算价"栏填写最终结算金额',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
            'wide': {
                'filename': 'payment_import_template_wide.csv',
                'headers': [
                    '合同编号或序号',
                    '结算价（元）',
                    '是否办理结算',
                ] + [f'{year}年{month}月' for year in range(BASE_YEAR, get_current_year() + 2) for month in range(1, 13)] + ['模板说明'],
                'notes': [
                    '【宽表格式】第1列填写合同编号或合同序号，后续月份列填写当期付款金额',
                    '【月份范围】已预设2019年1月至2025年12月共84个月份列，覆盖常用时间范围',
                    '【金额填写】每个月份列中填写当月的付款金额，单位为元，仅填写数字',
                    '【结算信息】如某合同已办理结算，在"结算价"和"是否办理结算"列填写相应信息',
                    '【留空规则】无付款的月份留空即可，不影响导入',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
        },
        'supplier_eval': {
            'long': {
                'filename': 'supplier_eval_import_template_long.csv',
                'headers': [
                    '序号',
                    '合同编号',
                    '履约综合评价得分',
                    '末次评价得分',
                ] + [f'{year}年度评价得分' for year in range(BASE_YEAR, get_current_year() + 2)] + [
                    '第1次过程评价得分',
                    '第2次过程评价得分',
                    '备注',
                    '模板说明',
                ],
                'notes': [
                    '必填：序号*、合同编号*；其余可留空。',
                    '评价编号由系统基于"EVAL-<合同编码>-<序号>"规则自动生成。',
                    '分数范围0-100，可保留1-2位小数；留空不导入该项。',
                    '可选：按年动态列（如"2024年度评价得分"）会自动识别，无需固定年份。',
                    '可选：过程评价列（如"第1次过程评价得分"、"第2次过程评价得分"）会自动识别。',
                    '模板说明列可删除，不影响导入。'
                ],
            },
        },
        'evaluation': {
            'long': {
                'filename': 'supplier_evaluation_import_template_long.csv',
                'headers': [
                    '项目编码',
                    '序号',
                    '评价编号',
                    '关联合同编号',
                    '供应商名称',
                    '评价日期区间',
                    '评价人员',
                    '评分',
                    '评价类型',
                    '模板说明',
                ],
                'notes': [
                    '【必填字段】评价编号*、关联合同编号*、供应商名称*（标记*号的为必填字段，不能为空）',
                    '【编码规则】评价编号须遵守编号格式限制（禁止 / 等特殊字符），推荐格式：HT2024-001-PJ01',
                    '【合同关联】关联合同编号必须填写系统中已存在的合同编号',
                    '【评分范围】评分范围为 0-100 之间的数字，可带小数（如：85.5），可留空',
                    '【评价类型】建议填写：履约过程评价、末次评价、阶段性评价等',
                    '【日期区间】评价日期区间格式示例：2024年1-6月、2024年上半年、2024Q1等',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
            'wide': {
                'filename': 'supplier_evaluation_import_template_wide.csv',
                'headers': [
                    '关联合同编号',
                    '供应商名称',
                ] + [f'{year}年{half}' for year in range(BASE_YEAR, get_current_year() + 2) for half in ['上半年', '下半年']] + ['模板说明'],
                'notes': [
                    '【宽表格式】第1列填写合同编号，第2列填写供应商名称',
                    '【评价周期】已预设2019年至2025年，每年上下半年共14个评价周期列',
                    '【评分填写】每个周期列中填写对应时期的评分（0-100），可保留一位或两位小数',
                    '【留空规则】如某时期暂无评价，对应单元格可留空',
                    '【说明】本模板说明行可保留或删除，不影响导入。导入时系统会自动跳过说明行。',
                ],
            },
        },
    }


class TemplateGenerator:
    """模板生成器"""
    
    def __init__(self, config_path: str):
        """
        初始化生成器
        
        Args:
            config_path: 配置文件路径
        """
        self.config_path = Path(config_path)
        self.config = self._load_config()
        self.model = self._get_model()
        
    def _load_config(self) -> Dict[str, Any]:
        """加载配置文件"""
        with open(self.config_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    
    def _get_model(self):
        """获取 Django 模型类"""
        module_name = self.config['metadata']['module']
        model_name = self.config['metadata']['model']
        return apps.get_model(module_name, model_name)
    
    def _get_choices_text(self, field_config: Dict) -> str:
        """获取选项文本"""
        if not field_config.get('choices_from_model'):
            return ""
        
        field_name = field_config['field']
        # 处理关联字段（如 project__project_name）
        if '__' in field_name:
            field_name = field_name.split('__')[0]
        
        try:
            model_field = self.model._meta.get_field(field_name)
            
            if hasattr(model_field, 'choices') and model_field.choices:
                choices = [display for value, display in model_field.choices]
                return "、".join(choices)
        except Exception as e:
            print(f"警告: 无法获取字段 {field_name} 的choices: {e}")
        
        return ""
    
    def _format_instructions(self) -> str:
        """
        格式化说明文本
        优先从 helptext 配置获取，如果不存在则使用模板配置
        """
        module_name = self.config['metadata']['module']
        
        # 尝试从 helptext 配置获取导入说明
        try:
            instructions = get_message('import', f'{module_name}_template')
            if instructions:
                return instructions
        except:
            pass
        
        # 降级到模板配置中的说明
        instructions = self.config['instructions']['content']
        
        # 替换占位符
        for field in self.config['fields']:
            if field.get('choices_from_model'):
                field_name = field['field'].split('__')[0] if '__' in field['field'] else field['field']
                placeholder = f"{{{field_name}_choices}}"
                choices_text = self._get_choices_text(field)
                if choices_text:
                    instructions = instructions.replace(placeholder, choices_text)
        
        return instructions
    
    def generate(self, output_path: str, year: Optional[int] = None) -> str:
        """
        生成模板文件
        
        Args:
            output_path: 输出目录
            year: 年份（用于文件名）
        
        Returns:
            生成的文件路径
        """
        if year is None:
            year = get_current_year()
        
        # 生成文件名
        filename_template = self.config['file']['name_template']
        filename = filename_template.format(year=year)
        filepath = Path(output_path) / filename
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = self.config['file']['sheet_name']
        
        # 样式定义
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        instruction_font = Font(size=10, color="FF0000")
        instruction_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入说明行
        instructions_row = self.config['instructions']['row']
        ws.merge_cells(start_row=instructions_row, start_column=1, 
                      end_row=instructions_row, end_column=len(self.config['fields']))
        cell = ws.cell(row=instructions_row, column=1)
        cell.value = self._format_instructions()
        cell.font = instruction_font
        cell.alignment = instruction_alignment
        ws.row_dimensions[instructions_row].height = 80
        
        # 写入表头
        header_row = instructions_row + 1
        for col_idx, field in enumerate(self.config['fields'], start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            
            # 字段名（必填字段标记*）
            field_name = field['name']
            if field.get('required'):
                field_name += "*"
            
            cell.value = field_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # 设置列宽
            ws.column_dimensions[cell.column_letter].width = 15
            
            # 添加批注（帮助文本）
            if field.get('help_text'):
                comment = Comment(field['help_text'], "系统")
                cell.comment = comment
        
        # 保存文件
        wb.save(filepath)
        return str(filepath)


def generate_all_templates(output_dir: str, year: Optional[int] = None) -> List[str]:
    """
    生成所有模板
    
    Args:
        output_dir: 输出目录
        year: 年份
    
    Returns:
        生成的文件路径列表
    """
    templates_dir = Path(__file__).parent / 'import_templates'
    
    # 确保输出目录存在
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    generated_files = []
    
    for config_file in templates_dir.glob('*.yml'):
        print(f"正在生成模板: {config_file.name}")
        try:
            generator = TemplateGenerator(config_file)
            filepath = generator.generate(output_dir, year)
            print(f"  ✓ 已生成: {filepath}")
            generated_files.append(filepath)
        except Exception as e:
            print(f"  ✗ 生成失败: {e}")
            import traceback
            traceback.print_exc()
    
    return generated_files


def generate_template_by_module(module: str, output_dir: str, year: Optional[int] = None) -> Optional[str]:
    """
    根据模块名生成模板
    
    Args:
        module: 模块名（procurement, contract, payment, supplier_eval）
        output_dir: 输出目录
        year: 年份
    
    Returns:
        生成的文件路径，失败返回 None
    """
    templates_dir = Path(__file__).parent / 'import_templates'
    config_file = templates_dir / f'{module}.yml'
    
    if not config_file.exists():
        print(f"错误: 模板配置不存在: {config_file}")
        return None
    
    # 确保输出目录存在
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    try:
        generator = TemplateGenerator(config_file)
        filepath = generator.generate(output_dir, year)
        print(f"✓ 已生成模板: {filepath}")
        return filepath
    except Exception as e:
        print(f"✗ 生成模板失败: {e}")
        import traceback
        traceback.print_exc()
        return None